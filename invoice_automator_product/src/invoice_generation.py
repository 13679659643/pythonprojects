# -*- coding: utf-8 -*-
# @Time    : 2024/9/4 17:39
# @Author  : Night
# @File    : invoice_generation.py
# @Description: 发票模板生成
"""
读取源数据->各个表互相映射->生成固定模板
"""
import json
import os
import time

import openpyxl
import requests
from digiCore.message.config import MSG_URL
from openpyxl.styles import PatternFill, Font
from settings import dim_invoice_amz_tracking_code, dim_invoice_amz_tracking_code_field_list, dingding_group_api, \
    redis_fba_info_key
from loguru import logger
from digiCore.db.redis.core import RedisDao
from digiCore.message.to_user import ToUser
from src.crawler_lx_fba_shipment_carton import CrawlerLXFbaShipmentCarton
from utils.tool import upload_media
from digiCore.message.to_webhook import ToWebhook
from settings import dingding_api, redis_img_key, appType, systemToken
from openpyxl.drawing.image import Image
from io import BytesIO
from src.yida_download_attachments import YidaDownloadAttachments
from src.data_processing import DataProcessing
from PIL import Image as PILImage


class InvoiceGeneration(DataProcessing):
    def __init__(self, shipment_id, amz_code, redis_run_sign=''):
        super().__init__()
        self.shipment_id = shipment_id
        self.amz_code = amz_code
        self.redis_client = RedisDao()
        self.dd_webhook = ToWebhook(dingding_api, atUserIds=['16765099110841587'])
        self.run_sign = redis_run_sign
        self.yd_client = YidaDownloadAttachments(appType=appType,
                                                 systemToken=systemToken)

    def insert_image_to_excel(self, ws, cell, image_data):
        """
        插入图片到excel 表中
        :param ws:
        :param cell:
        :param image_data:
        :return:
        """
        image = Image(image_data)
        ws.add_image(image, cell)

    def check_declaration_info(self, spu: str, color: str, form_value: dict):
        """
        检查报关 配置信息 填写
        材质
        针对材质的填写
        注意：金属头  关税影响
            飞织，统一改为织面 侵权风险
        :param spu: 原始货号
        :param color: 原始颜色
        :param form_value: 报关信息
        :return:
        """
        if not form_value['customs_declaration_material']:
            message = f'监控 <font color="red">{spu} {color}</font> 未配对报关材质'
            self.dd_webhook.send_markdown_message(message, '监控通知')

    def replace_none_with_empty(self, input_dict):
        """
        将字典中所有值为 'None' 的键替换为空字符串 ''。
        """
        return {k: ('' if v == 'None' else v) for k, v in input_dict.items()}

    def get_form_attachments(self, box_list: list[dict]):
        """
        获取表单数据
        :return:
        """
        # 使用集合去重
        original_spu_set = {box['original_spu'] for box in box_list}
        original_color_set = {box['original_color'] for box in box_list}

        # 将列表元素转换为 SQL 查询中所需的字符串格式
        spu_tuple = "(" + ",".join(f"'{spu}'" for spu in original_spu_set) + ")"
        color_tuple = "(" + ",".join(f"'{color}'" for color in original_color_set) + ")"
        sql = f"""
        SELECT original_spu,original_color,image,en_category,customs_declaration_material,purpose
        FROM dim_prod.dim_cd_prd_yida_product_info_i_manual where original_spu in {spu_tuple}
        and original_color in {color_tuple}
        """
        original_data = self.tidb.query_list(sql)
        form_dict = {}
        for row in original_data:
            form_key = row['original_spu'] + row['original_color']
            image_str = row['image']
            image_path = ''
            if image_str != 'None':
                image_obj = json.loads(image_str)
                image_path = image_obj[0]['downloadUrl']
            form_value = self.replace_none_with_empty({
                'en_category': row['en_category'],  # 品类(英文)
                'customs_declaration_material': row['customs_declaration_material'],  # 报关材质
                'purpose': row['purpose'],  # 报关用途
                'img_path': image_path  # 图片地址
            })
            # 检查报关材质是否填写
            self.check_declaration_info(row['original_spu'], row['original_color'], form_value)

            form_dict[form_key] = form_value
        return form_dict

    def matching_customs_code_rules(self, form_code: dict, original_spu: str, category: str):
        """
        国外海关编码
        第一种：优先通过货号配对报关编码前6位，后4位补齐0
        第二种：货号配对为空的通过鞋类划分
                1.安全鞋：6401100000 （后面五个零）
                2.休闲鞋/运动鞋/厨师鞋：6404110000（后面四个零）
                3.骑行鞋：6402190090
                4.高尔夫鞋：6402190000
        :param form_code: 商品编码配对表
        :param original_spu: 货号
        :param category: 鞋类
        :return:
        """
        customs_code = form_code.get(original_spu)
        if not customs_code:
            if category == '安全鞋':
                customs_code = '6401100000'
            elif category in ['休闲鞋', '运动鞋', '厨师鞋']:
                customs_code = '6404110000'
            elif category == '骑行鞋':
                customs_code = '6402190090'
            elif category == '高尔夫鞋':
                customs_code = '6402190000'
        else:
            customs_code = customs_code[:6] + '0000'
        return customs_code

    def matching_country_declaration_amt(self, country: str, region: str):
        """
        英国渠道  申报金额 6美元
        欧洲（不含英国）和美国和加拿大渠道  申报金额 8美元
        英国假链接定价范围 20.99-23.8
        欧洲假链接定价范围 19.99-22.99
        申报的链接需与派送国一致，例如法国，那就是法国亚马逊的链接
        :return:
        """
        amt = 8
        if country == '英国':
            amt = 6
        elif (region == '欧洲' and country != '英国') or country in ['英国', '加拿大']:
            amt = 8
        return amt

    def image_exists(self, from_data: dict, row_key: str, ws, cell):
        """
        判断图片是否存在
        :param from_data:  表单
        :param row_key: 图片名
        :param ws: excel对象
        :param cell: 列
        :return:
        """
        img_file_name = row_key + '.png'
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_dir = os.path.join(script_dir, '../images')
        img_file_path = os.path.join(output_dir, img_file_name)
        with self.redis_client.conn as redis_conn:
            data = from_data.get(row_key, {})
            if not data:
                ws[cell] = ''
                return
            img_path = data.get('img_path')
            if not img_path:
                ws[cell] = ''
                return
            if redis_conn.exists(redis_img_key + row_key):
                # 如果存在，则直接从文件系统加载图片路径
                with open(img_file_path, 'rb') as fp:
                    img_data = fp.read()
            else:
                img_url = self.yd_client.get_attachment_download_url(img_path)
                img_data = self.yd_client.get_attachment_io(img_url)
                # 下载图片到文件下
                with open(img_file_path, 'wb') as fp:
                    fp.write(img_data)
                # 将键存入到redis中
                redis_conn.set(redis_img_key + row_key, "1")
            img = PILImage.open(BytesIO(img_data))
            # 设置目标宽度和高度
            target_width = 150
            target_height = 150

            # 计算缩放比例
            width_ratio = target_width / img.width
            height_ratio = target_height / img.height
            scale_ratio = min(width_ratio, height_ratio)

            # 计算新的尺寸
            new_width = int(img.width * scale_ratio)
            new_height = int(img.height * scale_ratio)

            img = img.resize((new_width, new_height), PILImage.LANCZOS)  # 调整图片大小
            img_byte_arr = BytesIO()
            img.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)

            # 将调整后的图片插入到 Excel 中
            self.insert_image_to_excel(ws, cell, img_byte_arr)

    def generate_excel(self, data):
        """
        生成发票模板 Excel
        :param data: 从数据库中获取的数据
        :return:
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        # 固定A列内容
        fixed_a_column = [
            "运单号（FBA单号）：（必填）", "交货仓库：（必填）", "交货客户代码：（必填）", "走货线路：（必填）",
            "报关方式：（必填）", "目的地/国家：（中文必填）", "地址类型*（必填）", "收货地址/亚马逊仓库代码*",
            "收件公司（商业地址必填）：", "收件人（商业地址必填）：", "联系电话（商业地址必填）：", "收件邮编（商业地址必填）：",
            "省份/州（商业地址必填）：", "城市（商业地址必填）：", "收件地址（商业地址必填）：", "总件数（必填）"
        ]
        # 固定C列内容
        fixed_c_column = [
            "", "是否购买保险（选填）：", "清关方式（必填）：", "发件人公司（英文）：", "发件人地址（英文）：",
            "VAT和EORI注册公司（进口商）：", "VAT和EORI注册地址（进口商）：", "VAT号：*", "EORI号：*",
            "VAT公司联系人：", "VAT公司电话：", "备注："
        ]
        # 固定E列内容
        fixed_e_column = [
            "务必填写规范：（只做内容填写，不允许调整格式，不允许合并单元格，填写内容不允许有空格）",
            "注意：表格中蓝色内容为必填项，黄色内容为选填项",
            "备注：",
            "1、成套的产品需注明单套包含的件数以及品名填写到S18栏备注",
            "2、鞋类产品需要提供鞋面材质+鞋底材质",
            "3、纺织类，服装类产品需要提供成分说明（水洗标一致） 95% cotton + 5% spandex",
            "4、垫子类，纸盒支架类需提供规格（如尺寸和高度），请填写到S18栏备注",
            "5、申报币种必须匹配：英国（英镑GBP），欧洲（欧元EUR），美国、日本、加拿大、其他国家（美金USD）-也可参考表格2",
            "6、一箱货有多个产品，箱号重复填写即可",
            "",
            "",
            "如有合并清关或者合并报关，需要明确哪几票合并然后填写",
        ]
        # 固定第17行内容
        fixed_17_row = [
            "亚马逊FBA子单号/箱唛号（必填）", "Reference ID（亚马逊追踪编码）", "中文品名（必填）", "英文品名（必填）",
            "材质（中文填写）（必填）", "用途（中文填写）（必填）", "国外海关编码（必填）", "产品类型/属性（必填）",
            "单箱产品数量PCS（必填）", "箱数/件数CTN（必填）",
            "申报单价币种（必填）", "单个产品申报单价", "单个产品净重KG(必填)", "产品高清图片（必须缩在方框内）",
            "品牌（如实申报）", "品牌类型（如实申报）",
            "型号（如实申报）", "销售链接（填写链接的前后不能有空格）", "备注（产品的其他特殊说明）", "总产品数量",
            "总申报金额", "SKUNO", "实重",
            "长", "宽", "高"
        ]
        fixed_17_colors = [
            "00B0F0", "FFFF00", "00B0F0", "00B0F0", "00B0F0", "00B0F0", "00B0F0", "00B0F0", "00B0F0", "00B0F0",
            "00B0F0", "00B0F0", "00B0F0",
            "00B0F0", "FFFF00", "FFFF00", "FFFF00", "FFFF00", "FFFF00", "00B0F0", "00B0F0", "FFFF00",
            "FFFF00",
            "FFFF00", "FFFF00", "FFFF00"
        ]
        # 填充A列内容
        ws.column_dimensions['A'].width = 50
        for i, value in enumerate(fixed_a_column, start=1):
            cell = ws[f'A{i}']
            cell.value = value
            cell.font = Font(name='微软雅黑', size=12)
            cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

        # 填充B列内容
        ws.column_dimensions['B'].width = 30
        ws['B1'] = data['shipment_id']
        ws['B2'] = '泉州仓'
        ws['B3'] = ''
        ws['B4'] = data['logistics_channel']
        ws['B5'] = '客户自主报关'  # 目前写死
        ws['B6'] = data['zh_country']
        ws['B7'] = '亚马逊地址'
        ws['B8'] = data['ads'] + "\n" + data['country'] + ' ' + data['fba_number']
        for i in range(9, 16):
            ws[f'B{i}'] = ''
        for i in range(1, 8):
            ws[f'B{i}'].font = Font(name='微软雅黑', size=12)

        ws['B16'] = data['last_success_box_count']
        ws['B16'].font = Font(name='微软雅黑', size=26)
        # 填充C列值
        ws.column_dimensions['C'].width = 50
        for i, value in enumerate(fixed_c_column, start=1):
            cell = ws[f'C{i}']
            cell.value = value
            cell.font = Font(name='微软雅黑', size=12)
            if i in {2, 6, 10, 11, 12}:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            if i in {3, 4, 5, 8, 9}:
                cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        # 填充D列值
        ws.column_dimensions['D'].width = 30
        ws['D1'] = ''
        ws['D2'] = '否'
        ws['D3'] = data['clearance_type']
        ws['D4'] = data['address1']
        ws['D5'] = data['address2']
        ws['D6'] = ''
        ws['D7'] = ''
        ws['D8'] = data['vat']
        ws['D9'] = data['eori']
        ws['D10'] = ''
        ws['D11'] = ''
        ws['D12'] = ''
        for i in range(1, 12):
            ws[f'D{i}'].font = Font(name='微软雅黑', size=12)
        # 填充E列内容并设置背景颜色
        for i, value in enumerate(fixed_e_column, start=1):
            cell = ws[f'E{i}']
            cell.value = value
            cell.font = Font(name='微软雅黑', size=12)
            if i in {1, 2, 12}:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            if i in {4, 5, 6, 7, 8, 9}:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for i in range(1, 13):
            ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=15)  # 合并E列和F列

        # 固定 17行颜色
        for i, (value, color) in enumerate(zip(fixed_17_row, fixed_17_colors), start=1):
            cell = ws.cell(row=17, column=i, value=value)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(name='微软雅黑', size=12)
            if color == "FF0000":
                cell.font = Font(color="FF0000")

        box_data = self.get_box_list()  # 获取装箱数据
        if not box_data:
            message = f'监控 <font color="red">{data["shipment_id"]}</font> 货件装箱未采集'
            self.dd_webhook.send_markdown_message(message, '监控通知')
            self.redis_client.del_task_from_redis(self.run_sign)
            return
        form_data = self.get_form_attachments(box_data)  # 获取商品信息表
        form_code = self.get_customs_code()  # 获取海关编码
        ws.column_dimensions['N'].width = 25
        for i, row in enumerate(box_data, start=18):
            # ws.row_dimensions[i].height = 300
            ws.row_dimensions[i].height = 100
            box_key = row['original_spu'] + row['original_color']
            ws[f'A{i}'] = row['shipment_id'] + 'U00000' + row['box_id']  # FBA单号
            ws[f'B{i}'] = self.amz_code  # Reference ID
            ws[f'C{i}'] = row['category']
            ws[f'D{i}'] = form_data.get(box_key, {}).get('en_category', '')  # 品类(英文)
            ws[f'E{i}'] = form_data.get(box_key, {}).get('customs_declaration_material', '')  # 报关材质
            ws[f'F{i}'] = form_data.get(box_key, {}).get('purpose', '')  # 报关用途
            ws[f'G{i}'] = self.matching_customs_code_rules(form_code, row['original_spu'],
                                                           row['category'])  # 国外海关编码
            ws[f'H{i}'] = '普货'
            ws[f'I{i}'] = int(row['num'])
            ws[f'J{i}'] = 1
            ws[f'K{i}'] = 'USD'
            ws[f'L{i}'] = self.matching_country_declaration_amt(row['country'], row['region'])  # 申报金额
            ws[f'M{i}'] = 1
            self.image_exists(form_data, box_key, ws, f'N{i}')  # 判断图片是否存在
            ws[f'O{i}'] = row['brand']
            ws[f'P{i}'] = '境内自主品牌'
            ws[f'Q{i}'] = row['original_spu_color']
            ws[f'R{i}'] = row['sale_link']  # 销售链接
            ws[f'S{i}'] = ''
            ws[f'T{i}'] = int(row['num'])
            ws[f'U{i}'] = int(row['num']) * ws[f'L{i}'].value  # 总申报金额
            ws[f'V{i}'] = ''
            ws[f'W{i}'] = int(row['weight'])
            ws[f'X{i}'] = int(row['length'])
            ws[f'Y{i}'] = int(row['width'])
            ws[f'Z{i}'] = int(row['height'])
        # 获取当前脚本的绝对路径
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_dir = os.path.join(script_dir, '../output')
        file_name = data['logistics_provider_name'] + '-' + self.shipment_id + '-' + data['store_name'] + '.xlsx'
        file_input_path = os.path.join(output_dir, file_name)
        wb.save(file_input_path)
        # 发送消息到钉钉
        self.send_message(file_name, file_input_path, data['username'])

        # 同步到tidb 制作成功 更改状态
        self.sync_tidb(status=1)

    def send_message(self, file_name: str, file_input_path: str, user_name: str = '张致富'):
        """
        发送fba发票到指定的人
        :param file_name: 文件名
        :param file_input_path: 文件路径
        :param user_name: 指定人
        :return:
        """
        access_token = self.redis_client.get_dingding_access_token()
        media_id = upload_media(access_token, file_input_path)
        if not media_id:
            self.dd_webhook.send_text_message(f'监控 {file_name} 未发送到指定人')
            return
        user_sql = f"select userid from dim_prod.dim_dsd_yida_dd_employee_a_d where name='{user_name}'"
        user_data = self.tidb.query_one(user_sql)
        user_id = user_data['userid']
        t = ToUser([
            user_id
        ])
        t.send_file_message(media_id, file_name, fileType='xlsx')

    def check_declaration_is_exists(self, data):
        """
        检查 物流商推荐渠道 报关方式 清关方式 是否填写
        :return:
        """
        if not (data.get('logistics_channel') and data.get('declaration_type') and data.get('clearance_type')):
            shipment_id = data["shipment_id"]
            message = f'监控 <font color="red">{shipment_id}</font> 存在物流商报关信息为空情况'
            self.dd_webhook.send_markdown_message(message, '监控通知')
            return False
        return True

    def send_vat_message(self, store_name):
        """
        发送消息到指定人
        :param store_name: 店铺
        :return:
        """
        dd_webhook = ToWebhook(dingding_group_api, atUserIds=['16980414986533527'])
        markdown_message = f"![]({MSG_URL})" \
                           f"<h2>**<font color=\"#6FC7E1\">海关识别号VAT和EORI信息配对</font>**<h2>   \n" \
                           f"\n--- \n" \
                           f"\n<font color=\"#000000\">参与成员</font>： <font color=\"#444\">@16980414986533527</font> \n" \
                           f"\n--- \n" \
                           f"**<h2><font color=\"#6FC7E1\">对应店铺</font><h2>**   \n" \
                           f"{store_name}"
        dd_webhook.send_markdown_message(markdown_message, '监控通知')

    def split_string_with_last_dash(self, s, delimiter='-'):
        # 使用 rsplit 分割字符串，最大分割次数为 1
        parts = s.rsplit(delimiter, 1)

        if len(parts) > 1:
            # 分割后的结果有两部分，取第一部分加上分割符
            result = parts[0] + delimiter + 'FR'
            return result
        else:
            # 如果没有找到 '-'，则返回原字符串
            return s

    def check_vat_eori_is_exists(self, data):
        """
        检查 物流商推荐渠道 报关方式 清关方式 是否填写
        1. 英国和法国 EORI和VAT号  UK FR
        2. 其他国家 只有VAT号
        3. 美国 无税号 US
        3. NA 所有都无 EORI和VAT号
        :return:
        """
        store_name = data["store_name"]
        country = data['country']
        region = data['region']
        if country in ('UK', 'FR'):
            if not (data.get('vat') and data.get('eori')):
                self.send_vat_message(store_name)
                return data
            return data
        elif region != '北美':
            if not data.get('vat'):
                self.send_vat_message(store_name)
                return data
            elif region == '欧洲':
                store_name = self.split_string_with_last_dash(store_name)
                eori = self.get_store_eori(store_name)
                data['eori'] = eori
                return data
        return data

    def sync_logistics_provider(self):
        """
        更新货件信息
        :return:
        """
        url = 'http://110.191.179.224:8604/api/v1/logistics-provider-info-sync/schedule/syna_local_shipment_weight_data'
        response = requests.get(url)
        logger.info('------物流渠道信息更新-------')

    def check_shipment_id(self):
        """
        检查shipment_id 是否存在
        :return:
        """
        delivery_sql = f"""
        SELECT shipment_id FROM ods_prod.ods_scg_wld_me_shipment_delivery_record_i_h where shipment_id = '{self.shipment_id}'
        """
        delivery_id = self.tidb.query_one(delivery_sql)
        if not delivery_id:
            with self.redis_client.conn as redis_conn:
                if not redis_conn.exists(redis_fba_info_key):
                    redis_conn.set(redis_fba_info_key, "1")
                    self.sync_logistics_provider()
                    time.sleep(8)
                    redis_conn.delete(redis_fba_info_key)
                else:
                    logger.info('------物流渠道信息已经在更新-------')
                    time.sleep(8)
        sql = f"""
        SELECT shipment_id FROM dwd_prod.dwd_scg_wld_lx_fba_shipment_box_detail_a_d WHERE shipment_id = '{self.shipment_id}'
        """
        ship_id = self.tidb.query_one(sql)
        sql_item = f"""
                SELECT shipment_id FROM  dwd_prod.dwd_scg_wld_lx_fba_shipment_box_detail_item_i_d WHERE shipment_id = '{self.shipment_id}'
                """
        ship_item_id = self.tidb.query_one(sql_item)
        if not ship_id or not ship_item_id:
            cfc = CrawlerLXFbaShipmentCarton(self.shipment_id)
            cfc.run()

    def sync_tidb(self, status=0):
        """
        数据通过易搭点击 同步到dim层级
        :return:
        """
        data_list = [{"shipment_id": self.shipment_id, "amazon_tracking_code": self.amz_code, "status": status}]
        self.tidb.insert_data(dim_invoice_amz_tracking_code, dim_invoice_amz_tracking_code_field_list, data_list)
        logger.info(f'最新的亚马逊追踪编码{self.shipment_id}同步！')

    def main(self):
        # 检查装箱数据是否存在
        self.check_shipment_id()
        # 获取fba信息
        fba_data = self.get_fba_shipment_packed()
        # 检测VAT和EORI是否为空
        new_fba_data = self.check_vat_eori_is_exists(fba_data)
        # 检测物流信息是否为空
        status = self.check_declaration_is_exists(new_fba_data)
        if not status:
            self.redis_client.del_task_from_redis(self.run_sign)
            return
        self.generate_excel(new_fba_data)
        # 删除运行标识符
        self.redis_client.del_task_from_redis(self.run_sign)


if __name__ == "__main__":
    invo = InvoiceGeneration('FBA18HD5TKY9', '14AKHLDW')
    invo.main()
